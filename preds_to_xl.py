import ast

from openpyxl import Workbook
from openpyxl.compat import range

from utils import mapread
from nlp_utils import sort_by_word


def results_spreadsheet(task_name, word_preds, word_labels):
    wb = Workbook()
    ws = wb.active
    ws.title = "%s-reverse-img-preds" % task_name

    # header
    ws.cell(1, 1, 'word')
    ws.cell(1, 2, 'google translate label')
    ws.cell(1, 3, 'wiktionary label')
    for i in range(20):
        ws.cell(1, 4 + i, 'pred %s' % (i + 1))

    for i, ((word, preds), label_dict) in enumerate(list(zip(word_preds, word_labels))):
        pos_google = label_dict['pos_list'][0] if label_dict['pos_list'] else ''
        wik_pred = '%s (%s)' % (label_dict['pos_and_words'][0][1][0],
                                label_dict['pos_and_words'][0][0]) \
            if label_dict['pos_and_words'] else ''

        ws.cell(2 + i, 1, word)
        ws.cell(2 + i, 2, '%s (%s)' % (label_dict['t'], pos_google))
        ws.cell(2 + i, 3, wik_pred)

        for j, pred in enumerate(preds):
            ws.cell(2 + i, 4 + j, pred)

    wb.save(filename='%s_results.xlsx' % task_name)


if __name__ == '__main__':
    task_name  = 'fr_to_en'
    label_name = 'fr_labels2.txt'

    word_pred_path = '/Users/Shmu/PycharmProjects/translate-with-image/' \
                     'reverse-img-preds/%s/all_words' % task_name
    labels_en = '/Users/Shmu/PycharmProjects/translate-with-image/' \
                'labels/%s' % label_name

    word_preds, word_labels = [], []

    for word_dict in mapread(labels_en, ast.literal_eval):
        word = word_dict['w']
        path = osp.join(word_pred_path, word, 'preds.txt')
        try: io = open(path)
        except IOError as err: print(err)
        else:
            with io:
                word_labels.append(word_dict)
                word_preds.append((word, [p.strip() for p in io]))

    results_spreadsheet(task_name, word_preds, word_labels)